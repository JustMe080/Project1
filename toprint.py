import os
import win32print
import win32api
import openpyxl
from openpyxl import load_workbook

# Set printer to use long bond paper
def set_printer_to_long_bond(printer_name):
    try:
        printer_handle = win32print.OpenPrinter(printer_name)
        printer_info = win32print.GetPrinter(printer_handle, 2)
        dev_mode = printer_info['pDevMode']

        # Set paper size to long bond (8.5 x 13 inches)
        dev_mode.PaperSize = 14  # DMPAPER_FOLIO = 14
        win32print.SetPrinter(printer_handle, 2, printer_info, 0)
        win32print.ClosePrinter(printer_handle)

        print(f"Printer is set to long bond paper: {printer_name}")
    except Exception as e:
        print(f"Error setting printer paper size: {e}")

# Print all sheets in a single Excel file
def print_all_sheets(file_path, printer_name):
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            # Save the sheet as a temporary file
            temp_file = f"{file_path}_{sheet_name}.xlsx"
            new_workbook = openpyxl.Workbook()
            sheet = workbook[sheet_name]
            new_sheet = new_workbook.active
            new_sheet.title = sheet_name

            # Copy sheet content
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

            # Save temporary file
            new_workbook.save(temp_file)

            # Print the temporary sheet
            win32api.ShellExecute(
                0,
                "print",
                temp_file,
                None,
                ".",
                0
            )
            print(f"Printed sheet: {sheet_name} from {file_path}")

            # Delete temporary file
            os.remove(temp_file)
    except Exception as e:
        print(f"Failed to print {file_path}: {e}")

# Print all Excel files in a directory
def print_all_excel_files(directory, printer_name):
    set_printer_to_long_bond(printer_name)

    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith((".xls", ".xlsx")):
                file_path = os.path.join(root, file)
                print(f"Processing file: {file_path}")
                print_all_sheets(file_path, printer_name)

if __name__ == "__main__":
    # Folder containing Excel files
    folder_path = r"C:\project1\toprint"  # Your specified folder path
    printer_name = win32print.GetDefaultPrinter()  # Use the default printer or set manually

    print(f"Folder: {folder_path}")
    print(f"Printer: {printer_name}")

    if os.path.exists(folder_path):
        print_all_excel_files(folder_path, printer_name)
    else:
        print(f"Folder not found: {folder_path}")
